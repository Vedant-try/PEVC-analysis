import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --- Load and Clean the Data ---
@st.cache_data
def load_data():
    df = pd.read_excel("Deal Screening Data.xlsx", engine="openpyxl")
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df['Deal Value (USD mn)'] = pd.to_numeric(df['Deal Value (USD mn)'], errors='coerce')
    df['Buyer List'] = df['Buyer (s)'].fillna('').astype(str).apply(lambda x: [i.strip() for i in x.split(',') if i.strip()])
    df = df.explode('Buyer List')
    df['Buyer List'] = df['Buyer List'].str.strip()
    df['Formatted Date'] = df['Date'].dt.strftime('%d-%m-%Y')
    return df

df = load_data()

# --- Sidebar Filters ---
st.sidebar.title("🔍 Filters")

min_date = df['Date'].min().date()
max_date = df['Date'].max().date()
from_date = st.sidebar.date_input("From Date", min_date)
to_date = st.sidebar.date_input("To Date", max_date)

# Deal value input instead of slider
min_val = float(df['Deal Value (USD mn)'].min())
max_val = float(df['Deal Value (USD mn)'].max())

deal_value = st.sidebar.number_input("Minimum Deal Value (USD mn):", min_value=min_val, max_value=max_val, value=min_val)
st.sidebar.markdown(f"ℹ️ Min: **{min_val:.2f}**, Max: **{max_val:.2f}**")

# --- Filter Data ---
filtered_df = df[
    (df['Date'] >= pd.to_datetime(from_date)) &
    (df['Date'] <= pd.to_datetime(to_date)) &
    (df['Deal Value (USD mn)'] >= deal_value)
]

# --- Buyer Dropdown (Dynamic) ---
unique_buyers = sorted(filtered_df['Buyer List'].dropna().unique())
selected_buyers = st.sidebar.multiselect("Select Buyers", options=unique_buyers)

def prepare_excel_summary(data):
    summary = data.groupby('Buyer List').agg(
        Number_of_Deals=('Target Company Name', 'count'),
        First_Deal_Date=('Date', lambda x: x.min().strftime('%d-%m-%Y')),
        Last_Deal_Date=('Date', lambda x: x.max().strftime('%d-%m-%Y')),
        Min_Deal_Value=('Deal Value (USD mn)', lambda x: round(x.min(), 2)),
        Max_Deal_Value=('Deal Value (USD mn)', lambda x: round(x.max(), 2))
    ).reset_index()

    # Clean and deduplicate
    data = data.drop_duplicates(subset=['Buyer List', 'Target Company Name', 'Date', 'Deal Value (USD mn)'])

    # Expand deals for each buyer
    deal_dict = {}
    for buyer in data['Buyer List'].unique():
        buyer_df = data[data['Buyer List'] == buyer].sort_values('Date')
        deal_info = []
        for _, row in buyer_df.iterrows():
            deal_info.extend([
                row['Formatted Date'],
                row['Target Company Name'],
                round(row['Deal Value (USD mn)'], 2)
            ])
        deal_dict[buyer] = deal_info

    # Create expanded dataframe
    deals_df = pd.DataFrame.from_dict(deal_dict, orient='index')
    deals_df.index.name = "Buyer Name"
    deals_df.reset_index(inplace=True)

    # Dynamically create exact header count
    col_headers = ["Buyer Name"]
    num_deal_columns = deals_df.shape[1] - 1  # subtracting Buyer Name column
    num_deals = num_deal_columns // 3

    for i in range(1, num_deals + 1):
        col_headers += [f"Date {i}", f"Deal {i}", f"Deal Value {i} (in Mn USD)"]

    deals_df.columns = col_headers

    # Excel formatting
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary.to_excel(writer, sheet_name='Buyer Summary', index=False)
        deals_df.to_excel(writer, sheet_name='Buyer Deal Details', index=False)

        wb = writer.book
        ws1 = writer.sheets['Buyer Summary']
        ws2 = writer.sheets['Buyer Deal Details']

        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        for ws in [ws1, ws2]:
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
                for cell in col:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
                    if cell.row == 1:
                        cell.font = Font(bold=True)

        # Light coloring for deal groupings
        light_colors = ["#f2f2f2", "#e6f7ff", "#f9f9f9", "#eafbea"]
        for i in range(1, deals_df.shape[1]-1, 3):
            fill = PatternFill(start_color=light_colors[(i-1// 3) % len(light_colors)][1:], 
                               end_color=light_colors[(i-1 // 3) % len(light_colors)][1:], 
                               fill_type="solid")
            for row in ws2.iter_rows(min_row=2, min_col=i + 1, max_col=i + 3):
                for cell in row:
                    cell.fill = fill

    output.seek(0)
    return output


# --- Excel Download Button ---
import time  # make sure this is at the top of your file if not already

# --- Excel Download Button with Progress ---
st.sidebar.markdown("📤 Export Filtered Buyers")
if not filtered_df.empty:
    generate = st.sidebar.button("🛠 Generate Excel Output")

    if generate:
        with st.sidebar:
            st.markdown("🔄 Preparing your Excel file...")
            progress_bar = st.progress(0)

            # Simulate loading (you can tweak the speed if needed)
            for percent_complete in range(100):
                time.sleep(0.01)
                progress_bar.progress(percent_complete + 1)

            excel_output = prepare_excel_summary(filtered_df)

            st.success("✅ File is ready!")
            st.download_button(
                label="📥 Download Output.xlsx",
                data=excel_output,
                file_name="Output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# --- Main Display ---
st.title("💼 PE/VC Deal Explorer")

if not selected_buyers:
    st.info("Please select at least one buyer.")
else:
    for buyer in selected_buyers:
        st.markdown("---")
        st.subheader(f"📊 Summary for: {buyer}")
        buyer_df = filtered_df[filtered_df['Buyer List'] == buyer].drop_duplicates(
            subset=['Target Company Name', 'Date', 'Deal Value (USD mn)']).sort_values('Date')

        if buyer_df.empty:
            st.warning("No deals found.")
            continue

        st.markdown(f"""
        - **Total Deals**: {len(buyer_df)}
        - **First Investment**: {buyer_df['Date'].min().strftime('%d-%m-%Y')}
        - **Most Recent Investment**: {buyer_df['Date'].max().strftime('%d-%m-%Y')}
        - **Min Deal Value**: ${buyer_df['Deal Value (USD mn)'].min():,.2f} mn
        - **Max Deal Value**: ${buyer_df['Deal Value (USD mn)'].max():,.2f} mn
        """)

        st.write("### 💼 Investments")
        st.dataframe(buyer_df[['Formatted Date', 'Target Company Name', 'Deal Value (USD mn)', 'Deal Type']])

        st.write("### 🤝 Co-Investors")
        co_data = []
        for _, row in buyer_df.iterrows():
            co_investors = [b for b in str(row['Buyer (s)']).split(',') if b.strip() != buyer]
            co_data.append({
                'Date': row['Formatted Date'],
                'Target': row['Target Company Name'],
                'Co-Investors': ", ".join(co_investors),
                'Deal Type': row['Deal Type']
            })
        st.dataframe(pd.DataFrame(co_data))
